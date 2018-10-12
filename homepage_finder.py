import openpyxl as xl
import pprint as pp
import requests

def homepage_finder():
    #Load xlsx file into Python
    wb = xl.load_workbook('input.xlsx') #Open Input workbook
    journals = wb['journal_sheet'] #Open journals worksheet
    urls = wb["url_sheet"]
    num_rows = len(tuple(journals.rows))

    #journal_sheet column map
    #Col 0/A = journal name
    #Col 1/B = Publisher
    #Col 2/C = ISSN
    #Col 3/D = EISSN
    #Col 4/E = Country
    #Col 5/F = Language
    #Col 6/G = Category
    #Col 7/H = Submission Guidelines URL

    #Iterate through the url sheet to find the submission page URL outlines

    access_methods = {}
    publishers = []
    url_methods = []

    for row in urls.iter_rows(min_row=2, max_col=8, max_row=15):
        publishers.append(row[0].value)
        url_methods.append(row[7].value) 
        
    access_methods = dict(zip(publishers,url_methods))

    #Iterate through journals and replace None Values with empty strings
    for row in journals.iter_rows(min_row=2, max_col=7, max_row=num_rows):
        for col in row:
            if col.value == None:
                col.value = ""

    def get_url(row, source):
        #Alias varibles for readability
        
        journal = row[0]
        publisher = row[1]
        issn = row[2]
        eissn = row[3]
        target = row[7]
        
        #Check for each attribute in URL and replace with appropriate variable
        
        if "J_U_NAME" in source:
            #replace J_U_NAME with lowercase journal name and replace space with an underscore
            temp_name = journal.value.lower().replace(" ","_")
            target = source.replace("J_U_NAME", temp_name)
            source = target
        if "JNAME" in source:
            #Replace JNAME with lowercase journal name and replace space with hyphen
            temp_name = journal.value.lower().replace(" ", "-") 
            target = source.replace("JNAME",temp_name)
            source = target
        if "EISSN" in source:
            #Replace EISSN with actual EISSN, replacing space with hyphen
            temp_name = eissn.value.replace("-","")
            target = source.replace("EISSN", temp_name)
            source = target
        if "E_H_SSN" in source:
            #Replace EISSN with actual EISSN, keeping hyphens
            temp_name = eissn.value
            target = source.replace("E_H_SSN", temp_name)
            source = target
        if "ISSN" in source:
            #Replace ISSN with actual ISSN, replacing space with hyphen
            temp_name = issn.value.replace("-","")
            target = source.replace("ISSN", temp_name)
            source = target
        if "I_H_SSN" in source:
            #Replace ISSN with actual ISSN, keeping hyphens
            temp_name = issn.value
            target = source.replace("I_H_SSN", temp_name)
            source = target
        return target

    #Iterate over all rows, find the URL outline and return the submission outline Page

    for row in journals.iter_rows(max_col=9, min_row=2, max_row=num_rows):
      
        #Check if submission page has already been found
        if(row[7].value == None):
            #if not found, search for submission page
            for key, value in access_methods.items():
                if key == row[1].value:
                    try:
                        row[7].value = get_url(row, access_methods[row[1].value])#Pass row and url outline
                    except:
                        continue
                        
        #Check if home page has already been found
        try:
            found = row[8].value[0:4] == "ERROR"
            
        except:
            found = True
        
        if(found == False):
            #if not found, try each access method to find home page
            for key, value in access_methods.items():
                if(access_methods[key]!= None and access_methods[key][0:25] != "https://www.elsevier.com/"): #Exclude elsevier
                    candidate_url = get_url(row, access_methods[key])#Pass row and url outline
                    try:
                        r = requests.get(candidate_url) #Search for url
                    
                        if(r.status_code != 404): #If url does not return a 404, use that url
                            row[6].value = candidate_url
                            break
                    except:
                        continue
                    
            
