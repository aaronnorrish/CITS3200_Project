import openpyxl as xl
import pprint as pp
import urllib3
from bs4 import BeautifulSoup
import requests
import numpy as np
import time
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.chrome.options import Options as ChromeOptions
import datetime

# This program takes in a given xlsx file and finds the appropriate submission guidelines url
# and home page url for a given journal by scraping the web. This program takes along time to run
# as it needs to scrape many pages aswell as wait for arbitrary amounts of time to avoid IP blocking.
# For this reason, its ideal to run this program is small batches.

# This program may only be run at UWA as the university has special IP privilleges that allow it to 
# Scrape researchgate websites with out being blocked



def open_workbook(filename):
    #Load xlsx file into Python
    wb = xl.load_workbook('test.xlsx') #Open Input workbook
    journals = wb['journal_sheet'] #Open journals worksheet
    urls = wb["url_sheet"]
    num_rows = len(tuple(journals.rows))
    #journal_sheet column map
    #Col 0/A = journal name
    #Col 1/B = Publisher
    #Col 2/C = ISSN
    #Col 3/D = EISSN
    #Col 4/E = Country
    #Col 5/F = Language
    #Col 6/G = Category
    #Col 7/H = Submission Guidelines URL


# This function attempts to match a given url candidate to its actual url
# by parsing the candidate url and matching it with the appropriate data 
# to produce a searchable URL
# Example:
#     source = "www.wiley.com/ISSN"
#     issn = 1234-5678
#     target = "www.wiley.com/12345678"

def get_url(row, source):
    #Alias varibles for readability
    journal = row[0]
    publisher = row[1]
    issn = row[2]
    eissn = row[3]
    target = row[8]
    
    #Check for each attribute in URL and replace with appropriate variable
    
    if "J_U_NAME" in source:
        #replace J_U_NAME with lowercase journal name and replace space with an underscore
        temp_name = journal.value.lower().replace(" ","_")
        target = source.replace("J_U_NAME", temp_name)
        source = target
    if "JNAME" in source:
        #Replace JNAME with lowercase journal name and replace space with hyphen
        temp_name = journal.value.lower().replace(" ", "-") 
        target = source.replace("JNAME",temp_name)
        source = target
    if "EISSN" in source:
        #Replace EISSN with actual EISSN, replacing space with hyphen
        temp_name = eissn.value.replace("-","")
        target = source.replace("EISSN", temp_name)
        source = target
    if "E_H_SSN" in source:
        #Replace EISSN with actual EISSN, keeping hyphens
        temp_name = eissn.value
        target = source.replace("E_H_SSN", temp_name)
        source = target
    if "ISSN" in source:
        #Replace ISSN with actual ISSN, replacing space with hyphen
        temp_name = issn.value.replace("-","")
        target = source.replace("ISSN", temp_name)
        source = target
    if "I_H_SSN" in source:
        #Replace ISSN with actual ISSN, keeping hyphens
        temp_name = issn.value
        target = source.replace("I_H_SSN", temp_name)
        source = target

    return target

# Set up firefox and chrome drivers
# These need to be used to avoid blocking IP
def connectChrome():
    options = ChromeOptions()
    options.add_argument("--headless")
    chromeDriverPath = "chromedriver.exe"
    driver = webdriver.Chrome(chromeDriverPath, chrome_options=options)
    print("Chrome Headless Browser Invoked")
    return driver

def connectFirefox():
    options = FirefoxOptions()
    options.add_argument("--headless")
    driver = webdriver.Firefox(firefox_options=options)
    print("Firefox Headless Browser Invoked")
    return driver


# This function will open the input file and search researchgate using Selenium to scrape the website
# It opens and closes the driver and saves the file every 100 iterations incase of failure
# Change the checkpoint variable to run from a different row incase the program failed halfway
def home_page_finder():
    #Load xlsx file into Python
    wb = xl.load_workbook('test.xlsx') #Open Input workbook
    journals = wb['journal_sheet'] #Open journals worksheet
    urls = wb["url_sheet"]
    num_rows = len(tuple(journals.rows))
    #journal_sheet column map
    #Col 0/A = journal name
    #Col 1/B = Publisher
    #Col 2/C = ISSN
    #Col 3/D = EISSN
    #Col 4/E = Country
    #Col 5/F = Language
    #Col 6/G = Category
    #Col 7/H = Submission Guidelines URL

    home_page_url = "https://www.researchgate.net/journal/I_H_SSN_J_U_NAME"

    delays = ((np.random.rand(5))) # create random delay times of 0 to 3 secs
    check_point = 2 #Change this to start iterating from a different row
    switch = 1

    n_searches = 0
    #Iterate over all rows, find URL
    #TAKES VERY VERY LONG
    for row in journals.iter_rows(max_col=9, min_row=2, max_row=num_rows):
        print("Starting Search for journal row item:" + str(check_point))
        
        #Check if submission page already found
        if(row[7].value != None):
            #if found, skip
            print("Skipping, already have submission page")
            check_point +=1
            continue
        
        #Check if homepage is already found
        if(row[8].value != None):
            #if found skip
            print("Skipping, previously searched homepage")
            check_point +=1
            continue
            
        #Save every 100 iterations and swap drivers
        if ((n_searches % 100) == 0):
            date = str(datetime.datetime.now()).replace(' ','')[0:18].replace(':','')
            wb.save('./logs/homepage/output_'+date+'checkpoint_'+str(check_point)+'.xlsx') #Save every 100 iterations
            #Swap and Init new Driver
            if switch == 0:
                try:
                    driver.quit()
                    driver = connectChrome()
                except:
                    driver = connectChrome()
                switch = 1
            elif switch == 1:
                try:
                    driver.quit()
                    driver = connectFirefox()
                except:
                    driver = connectFirefox()
                switch = 0
            
        #Delay by random amount of time
        delay = np.random.choice(delays)
        time.sleep(delay)
        
        #Open and Parse url
        driver.get(get_url(row, home_page_url))
        page = driver.page_source
        soup = BeautifulSoup(page, 'html.parser')
        n_searches +=1
        
        #Check if Page is on Website
        if(soup.find('h1').contents[0][0:17] == 'Directory results'): #if true, we have been redirected to Researchgate directory
            #This means the journal is not on researchgate
            error_message = "ERROR: Journal Not on Research Gate"
            row[8].value = error_message
            print(error_message)
            
        else:
            website_div = soup.find("table", {"class":"table journal-full-info__table"}).tbody.find_all("th")[4].contents[0]

            if website_div == 'Website':
                try:
                    url = soup.find("a", {"class":"nova-e-link nova-e-link--color-blue nova-e-link--theme-bare"}).contents[0]
                    row[8].value = url 
                    print(url)
                except:
                    error_message = "ERROR: No URL On Researchgate"
                    row[8].value = error_message
                    print(error_message)

        check_point +=1
    wb.save('test.xlsx')

# Attempts to find submission guidelines pages by matching candidate urls to appropriate data
def submission_page_finder():
    #Load xlsx file into Python
    wb = xl.load_workbook('test.xlsx') #Open Input workbook
    journals = wb['journal_sheet'] #Open journals worksheet
    urls = wb["url_sheet"]
    num_rows = len(tuple(journals.rows))

    #journal_sheet column map
    #Col 0/A = journal name
    #Col 1/B = Publisher
    #Col 2/C = ISSN
    #Col 3/D = EISSN
    #Col 4/E = Country
    #Col 5/F = Language
    #Col 6/G = Category
    #Col 7/H = Submission Guidelines URL

    #Iterate through the url sheet to find the submission page URL outlines

    access_methods = {}
    publishers = []
    url_methods = []

    for row in urls.iter_rows(min_row=2, max_col=8, max_row=15):
        publishers.append(row[0].value)
        url_methods.append(row[7].value) 
        
    access_methods = dict(zip(publishers,url_methods))

    #Iterate through journals and replace None Values with empty strings
    for row in journals.iter_rows(min_row=2, max_col=7, max_row=num_rows):
        for col in row:
            if col.value == None:
                col.value = ""

    #Iterate over all rows, find the URL outline and return the submission outline Page

    for row in journals.iter_rows(max_col=9, min_row=1, max_row=num_rows):
      
        #Check if submission page has already been found
        if(row[7].value == None):
            #if not found, search for submission page
            for key, value in access_methods.items():
                if key == row[1].value:
                    try:
                        row[7].value = get_url(row, access_methods[row[1].value])#Pass row and url outline
                    except:
                        continue
                        
        #Check if home page has already been found
        try:
            found = row[8].value[0:4] == "ERROR"
            
        except:
            found = True
        
        if(found == False):
            #if not found, try each access method to find home page
            for key, value in access_methods.items():
                if(access_methods[key]!= None and access_methods[key][0:25] != "https://www.elsevier.com/"): #Exclude elsevier
                    candidate_url = get_url(row, access_methods[key])#Pass row and url outline
                    try:
                        r = requests.get(candidate_url) #Search for url
                    
                        if(r.status_code != 404): #If url does not return a 404, use that url
                            row[6].value = candidate_url
                            break
                    except:
                        continue
    try:
      driver.quit()
    except:
        driver = connectChrome()
    date = str(datetime.datetime.now()).replace(' ','')[0:18].replace(':','')
    wb.save('test.xlsx')
                        
                    
#comment/uncomment below functions to run appropriate functions

submission_page_finder()
home_page_finder()
                